from mcp.server.fastmcp import FastMCP
import openpyxl
import sys
import os

mcp = FastMCP("excel_writer")

workbook = None
workbook_path = None


@mcp.tool()
def open_excel(path: str) -> str:
    """
    Open an Excel file to read and write.
    Example user prompt:
      "open excel E:\\test.xlsx"
    """
    global workbook, workbook_path
    abs_path = os.path.abspath(path)

    if not os.path.exists(abs_path):
        return f"File not found: {abs_path}"

    workbook = openpyxl.load_workbook(abs_path)
    workbook_path = abs_path
    return f"Opened {abs_path} | Sheets: {workbook.sheetnames}"


@mcp.tool()
def write(sheet: str, cell: str, value: str) -> str:
    """
    Write a value to a cell.
    Example user prompt:
      "write 12 in B4"
      (Claude will figure out sheet automatically if needed)
    """
    global workbook
    if workbook is None:
        return "No workbook open. Say: open excel <path> first."

    if sheet not in workbook.sheetnames:
        return f"Sheet '{sheet}' not found. Sheets: {workbook.sheetnames}"

    ws = workbook[sheet]
    ws[cell] = value
    return f"Placed '{value}' into {sheet}!{cell}. Say 'save excel' to commit changes."


@mcp.tool()
def save_excel() -> str:
    """Save the currently open Excel workbook."""
    global workbook, workbook_path
    if workbook is None:
        return "No workbook to save."

    workbook.save(workbook_path)
    return f"Saved: {workbook_path}"


if __name__ == "__main__":
    print("Excel MCP Server starting...", file=sys.stderr)
    mcp.run(transport="stdio")
